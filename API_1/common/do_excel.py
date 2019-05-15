"""
1，完成do_excel类的封装支持读和写，理解Case类的设计
2，结合http_requests完成注册，登录，充值接口的请求
3，尝试引入unittest+ddt 来完成以上接口的测试用例"""
import openpyxl
from API_1.common import http_requests

class Case:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql = None

class DoExcel:
    def __init__(self,file_name,sheet_name):
        try:
            self.file_name=file_name
            self.sheet_name=sheet_name
            self.workbook = openpyxl.load_workbook(file_name)
            self.sheet=self.workbook[sheet_name]
        except Exception as e:
            print("case文件有误：{}".format(e))

    def get_case(self):    #读取文件
        max_row = self.sheet.max_row #获取最大行

        cases = []  #存放所有的测试用例
        for r in range(2,max_row+1):   #遍历行数
            # case = {}
            # case["case_id"] = self.sheet.cell(row=r,column=1).value
            # case["title"] = self.sheet.cell(row=r, column=2).value
            # cases.append(case)

            case = Case()  #实例化
            case.case_id = self.sheet.cell(row=r,column=1).value
            case.title =self.sheet.cell(row=r, column=2).value
            case.url = self.sheet.cell(row=r, column=3).value
            case.data = self.sheet.cell(row=r, column=4).value
            case.method = self.sheet.cell(row=r, column=5).value
            case.expected = self.sheet.cell(row=r, column=6).value
            case.sql = self.sheet.cell(row=r,column=9).value     #执行的sql语句
            cases.append(case)

            self.workbook.close()

        return cases #返回case列表

    def write_result(self,row,actual,result):  #固定行数，写入值
        sheet = self.workbook[self.sheet_name]  #确认sheet
        sheet.cell(row,7).value = actual
        sheet.cell(row,8).value = result

        self.workbook.save(filename=self.file_name)
        self.workbook.close()

if __name__ == '__main__':
    from API_1.common import contants
    do_excel = DoExcel(contants.api_case_file,sheet_name='login')
    cases = do_excel.get_case()  #调用do_excel类中的get_case()
    http_request = http_requests.HttpRequests2()   #调用http_requests中的HttpRequests类
    for case in cases:  #遍历cases
        # print(case.case_id)
        # print(case.method)
        # print(case.data)
        print(case.__dict__) #__dict__属性，把所有的case属性放入字典
        print(type(case.data))   #从case里面读出来的是str，但是我们传入要是字典
        resp = http_request.request(case.method,case.url,case.data)
        # print(resp.status_code)
        # print(resp.text)
        # print(resp)
        # resp_dict = resp.json()['code']
        # print("resp_dict是",resp_dict)
        # """把resp.text返回字典,在文本内容很多的时候，可以用来判断某个元素就可以了"""
        actual = resp.text   #
        if case.expected == actual:  #判断期望结果与实际结果是否一致,判断的是响应文本
            do_excel.write_result(case.case_id+1,actual,'PASS')  #case为文件cases里面的
        else:
            do_excel.write_result(case.case_id+1,actual,'FAIL')

